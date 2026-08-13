"""Serialize corrected transcription JSON as JSON, CSV or styled XLSX."""

from __future__ import annotations

from copy import copy
import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
import json
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .tables import ExportTable, build_table


ExportFormat = Literal["xlsx", "csv", "json"]

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="173772")
_YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF3CD")
_RED_FILL = PatternFill(fill_type="solid", fgColor="F8D7DA")
_RED_LEFT = Side(style="medium", color="DC3545")


@dataclass(frozen=True, slots=True)
class ExportedFile:
    content: bytes
    media_type: str
    extension: str


def _csv_bytes(table: ExportTable) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(table.headers)
    writer.writerows(table.rows)
    return output.getvalue().encode("utf-8")


def _style_header(worksheet, width: int) -> None:
    for cell in worksheet[1][:width]:
        cell.fill = _HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_warning_rows(worksheet, table: ExportTable) -> None:
    for row_number, warning in enumerate(table.warnings, start=2):
        if warning == "none":
            continue
        fill = _RED_FILL if warning == "red" else _YELLOW_FILL
        for cell in worksheet[row_number][: len(table.headers)]:
            cell.fill = fill
        if warning == "red":
            first = worksheet.cell(row=row_number, column=1)
            border = copy(first.border)
            first.border = Border(
                left=_RED_LEFT,
                right=border.right,
                top=border.top,
                bottom=border.bottom,
                diagonal=border.diagonal,
                diagonal_direction=border.diagonal_direction,
                diagonalUp=border.diagonalUp,
                diagonalDown=border.diagonalDown,
                outline=border.outline,
                vertical=border.vertical,
                horizontal=border.horizontal,
            )


def _size_columns(worksheet, table: ExportTable) -> None:
    for index, header in enumerate(table.headers, start=1):
        longest = len(str(header))
        for row in table.rows:
            longest = max(longest, len(str(row[index - 1])))
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            max(longest + 2, 10), 45
        )


def _xlsx_bytes(table: ExportTable) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Transcrição"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.append(table.headers)
    for row in table.rows:
        worksheet.append(row)
    _style_header(worksheet, len(table.headers))
    _style_warning_rows(worksheet, table)
    _size_columns(worksheet, table)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_transcription(
    tipo: str,
    formato: ExportFormat,
    value: dict[str, Any],
) -> ExportedFile:
    if formato == "json":
        content = json.dumps(value, ensure_ascii=False, separators=(",", ":")).encode(
            "utf-8"
        )
        return ExportedFile(content, "application/json", "json")

    table = build_table(tipo, value)
    if formato == "csv":
        return ExportedFile(_csv_bytes(table), "text/csv; charset=utf-8", "csv")
    if formato == "xlsx":
        return ExportedFile(
            _xlsx_bytes(table),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx",
        )
    raise ValueError("Formato de exportação desconhecido")
