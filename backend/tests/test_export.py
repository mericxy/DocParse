from __future__ import annotations

import csv
from io import BytesIO, StringIO
import json

from openpyxl import load_workbook
import pytest

from backend.app.services.export import AmbiguousPayrollExportError, export_transcription
from backend.app.services.export.tables import build_cartao_table, build_holerite_table


def _punch(kind: str, raw: str, normalized: str | None = None) -> dict[str, str]:
    return {"kind": kind, "time_raw": raw, "time_hhmm": normalized or raw}


def _field(label: str, value: str) -> dict[str, str]:
    return {"code": "", "label": label, "reference": "", "value": value}


def _holerite_page(
    page: int,
    year: str,
    month: str,
    *,
    fields: list[dict] | None = None,
    bases: list[dict] | None = None,
) -> dict:
    return {
        "page": page,
        "year": year,
        "month": month,
        "fields": fields or [],
        "bases": bases or [],
    }


def _rgb(cell) -> str:
    return cell.fill.fgColor.rgb or ""


def test_cartao_table_uses_max_punches_preserves_order_and_empty_day():
    value = {
        "pages": [
            {
                "page": 2,
                "days": [
                    {
                        "date_raw": "02/01/2025",
                        "punches": [
                            _punch("IN", "+08:00d", "08:00"),
                            _punch("OUT", "12:00"),
                            _punch("IN", "13:00"),
                        ],
                    },
                    {"date_raw": "01/01/2025", "punches": []},
                ],
            }
        ]
    }

    table = build_cartao_table(value)

    assert table.headers == ["Data", "Entrada 1", "Saída 1", "Entrada 2", "Saída 2"]
    assert table.rows == [
        ["02/01/2025", "08:00", "12:00", "13:00", ""],
        ["01/01/2025", "", "", "", ""],
    ]
    assert table.warnings == ["yellow", "red"]


def test_cartao_red_warning_wins_over_odd_or_uncertain_punch():
    value = {
        "pages": [
            {
                "page": 1,
                "days": [
                    {"date_raw": "01/01/2025", "punches": []},
                    {
                        "date_raw": "03/01/2025",
                        "punches": [_punch("IN", "0?:00")],
                    },
                ],
            }
        ]
    }
    assert build_cartao_table(value).warnings == ["none", "red"]


def test_cartao_day_only_sequence_is_checked_only_inside_each_physical_page():
    value = {
        "pages": [
            {
                "page": 1,
                "days": [
                    {"date_raw": "01", "punches": []},
                    {"date_raw": "02", "punches": []},
                    {"date_raw": "03", "punches": []},
                ],
            },
            {
                "page": 2,
                "days": [
                    {"date_raw": "30", "punches": []},
                    {"date_raw": "31", "punches": []},
                ],
            },
            {"page": 3, "days": [{"date_raw": "01", "punches": []}]},
        ]
    }

    assert build_cartao_table(value).warnings == [
        "none",
        "none",
        "none",
        "none",
        "none",
        "none",
    ]

    value["pages"][0]["days"][1]["date_raw"] = "03"
    assert build_cartao_table(value).warnings[:3] == ["none", "red", "red"]


def test_holerite_columns_follow_first_appearance_and_missing_field_is_empty():
    value = {
        "pages": [
            _holerite_page(
                1,
                "2024",
                "12",
                fields=[_field("Verba B", "20,00"), _field("Verba A", "10,00")],
                bases=[{"label": "Base INSS", "value": "30,00"}],
            ),
            _holerite_page(2, "2025", "01", fields=[_field("Verba A", "11,00")]),
        ]
    }

    table = build_holerite_table(value)

    assert table.headers == ["Pág.", "Mês", "Ano", "Verba B", "Verba A"]
    assert "Base INSS" not in table.headers
    assert table.rows == [
        [1, "12", "2024", "20,00", "10,00"],
        [2, "01", "2025", "", "11,00"],
    ]
    assert table.warnings == ["none", "none"]


def test_holerite_warnings_ignore_unreadable_competence_and_red_wins():
    value = {
        "pages": [
            _holerite_page(1, "2024", "12", fields=[_field("A", "1,00")]),
            _holerite_page(2, "2025", "??"),
            _holerite_page(3, "2025", "01", fields=[_field("A", "1,00")]),
            _holerite_page(4, "2025", "03", fields=[_field("A", "1,?0")]),
        ]
    }

    table = build_holerite_table(value)

    assert table.warnings == ["none", "yellow", "none", "red"]


def test_holerite_repeated_competences_are_additional_blocks_not_sequence_breaks():
    value = {
        "pages": [
            _holerite_page(1, "2024", "08", fields=[_field("A", "1,00")]),
            _holerite_page(1, "2024", "08", fields=[_field("B", "2,00")]),
            _holerite_page(2, "2024", "09", fields=[_field("A", "3,00")]),
            _holerite_page(2, "2024", "09", fields=[_field("B", "4,00")]),
            _holerite_page(3, "2024", "10", fields=[_field("A", "5,00")]),
            _holerite_page(4, "2024", "12", fields=[_field("A", "6,00")]),
            _holerite_page(4, "2024", "12", fields=[_field("B", "7,00")]),
            _holerite_page(5, "2025", "01", fields=[_field("A", "8,00")]),
        ]
    }

    assert build_holerite_table(value).warnings == [
        "none",
        "none",
        "none",
        "none",
        "none",
        "red",
        "none",
        "none",
    ]


def test_holerite_csv_and_xlsx_refuse_duplicate_labels_in_one_logical_row():
    value = {
        "pages": [
            _holerite_page(
                2,
                "2017",
                "12",
                fields=[
                    _field("419 13º. Adto Desc", "100,00"),
                    _field("419 13º. Adto Desc", "9.999,99"),
                ],
            )
        ]
    }

    for formato in ("csv", "xlsx"):
        with pytest.raises(AmbiguousPayrollExportError, match="label repetido"):
            export_transcription("holerite", formato, value)

    exported = export_transcription("holerite", "json", value)
    assert json.loads(exported.content)["pages"][0]["fields"] == value["pages"][0]["fields"]


def test_xlsx_uses_required_header_and_warning_styles():
    value = {
        "pages": [
            {
                "page": 1,
                "days": [
                    {"date_raw": "01/01/2025", "punches": [_punch("IN", "08:00")]},
                    {
                        "date_raw": "03/01/2025",
                        "punches": [_punch("IN", "0?:00")],
                    },
                ],
            }
        ]
    }
    exported = export_transcription("cartao-ponto", "xlsx", value)
    workbook = load_workbook(BytesIO(exported.content))
    worksheet = workbook.active

    assert worksheet["A1"].value == "Data"
    assert worksheet["A1"].font.bold is True
    assert worksheet["A1"].font.color.rgb.endswith("FFFFFF")
    assert _rgb(worksheet["A1"]).endswith("173772")
    assert _rgb(worksheet["A2"]).endswith("FFF3CD")
    assert _rgb(worksheet["A3"]).endswith("F8D7DA")
    assert worksheet["A3"].border.left.color.rgb.endswith("DC3545")


def test_json_csv_and_xlsx_exports_contain_real_data_for_both_types(
    cartao_value, holerite_value
):
    for tipo, value, expected_header in (
        ("cartao-ponto", cartao_value, "Data"),
        ("holerite", holerite_value, "Pág."),
    ):
        json_export = export_transcription(tipo, "json", value)
        assert json.loads(json_export.content.decode("utf-8")) == value

        csv_export = export_transcription(tipo, "csv", value)
        csv_rows = list(csv.reader(StringIO(csv_export.content.decode("utf-8"))))
        assert csv_rows[0][0] == expected_header
        assert len(csv_rows) == 2

        xlsx_export = export_transcription(tipo, "xlsx", value)
        workbook = load_workbook(BytesIO(xlsx_export.content), read_only=True)
        worksheet = workbook.active
        assert worksheet.cell(1, 1).value == expected_header
        assert worksheet.max_row == 2
