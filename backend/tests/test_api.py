from __future__ import annotations

import csv
from datetime import datetime, UTC
from io import BytesIO, StringIO

from openpyxl import load_workbook
from sqlalchemy import select

from backend.app.db.models import Transcription


def _now() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)


def _insert_job(client, *, job_id: str, tipo: str, status: str, value=None, erro=None):
    with client.app.state.session_factory() as session:
        session.add(
            Transcription(
                id=job_id,
                tipo=tipo,
                status=status,
                erro=erro,
                value=value,
                file_path=str(client.app.state.settings.upload_dir / f"{job_id}.pdf"),
                created_at=_now(),
                started_at=None,
                finished_at=None,
            )
        )
        session.commit()


def test_post_valid_pdf_returns_202_without_calling_extractor(
    client, valid_pdf_bytes, monkeypatch
):
    def forbidden_extractor(_path):
        raise AssertionError("o extrator não pode rodar no POST")

    monkeypatch.setattr(
        "backend.app.services.extraction.extract_cartao_ponto", forbidden_extractor
    )
    response = client.post(
        "/api/transcricoes",
        data={"tipo": "cartao-ponto"},
        files={"arquivo": ("nome-com-pii.pdf", valid_pdf_bytes, "application/pdf")},
    )

    assert response.status_code == 202
    job_id = response.json()["id"]
    with client.app.state.session_factory() as session:
        job = session.scalar(select(Transcription).where(Transcription.id == job_id))
        assert job is not None
        assert job.status == "processando"
        assert job.started_at is None
        assert job.file_path.endswith(f"{job_id}.pdf")
        assert "nome-com-pii" not in job.file_path


def test_post_rejects_invalid_type_oversize_and_invalid_pdfs(client, valid_pdf_bytes):
    invalid_type = client.post(
        "/api/transcricoes",
        data={"tipo": "outro"},
        files={"arquivo": ("a.pdf", valid_pdf_bytes, "application/pdf")},
    )
    assert invalid_type.status_code == 422

    oversize = client.post(
        "/api/transcricoes",
        data={"tipo": "holerite"},
        files={"arquivo": ("a.pdf", b"%PDF-" + b"x" * (1024 * 1024), "application/pdf")},
    )
    assert oversize.status_code == 413

    fake = client.post(
        "/api/transcricoes",
        data={"tipo": "holerite"},
        files={"arquivo": ("a.pdf", b"texto renomeado", "application/pdf")},
    )
    assert fake.status_code == 422

    corrupt = client.post(
        "/api/transcricoes",
        data={"tipo": "holerite"},
        files={"arquivo": ("a.pdf", b"%PDF-1.7\ncorrompido", "application/pdf")},
    )
    assert corrupt.status_code == 422
    with client.app.state.session_factory() as session:
        assert session.scalar(select(Transcription.id)) is None
    assert list(client.app.state.settings.upload_dir.iterdir()) == []


def test_get_exposes_only_public_state_for_all_statuses(client, cartao_value):
    missing = client.get("/api/transcricoes/inexistente")
    assert missing.status_code == 404

    _insert_job(client, job_id="pending", tipo="cartao-ponto", status="processando")
    _insert_job(
        client,
        job_id="done",
        tipo="cartao-ponto",
        status="concluido",
        value=cartao_value,
    )
    _insert_job(
        client,
        job_id="failed",
        tipo="cartao-ponto",
        status="erro",
        erro="Falha legível.",
    )

    assert client.get("/api/transcricoes/pending").json() == {
        "id": "pending",
        "tipo": "cartao-ponto",
        "status": "processando",
        "erro": None,
        "value": None,
    }
    assert client.get("/api/transcricoes/done").json()["value"] == cartao_value
    assert client.get("/api/transcricoes/failed").json() == {
        "id": "failed",
        "tipo": "cartao-ponto",
        "status": "erro",
        "erro": "Falha legível.",
        "value": None,
    }


def test_put_validates_by_type_and_only_replaces_completed_value(
    client, cartao_value
):
    _insert_job(client, job_id="done", tipo="cartao-ponto", status="concluido", value=cartao_value)
    _insert_job(client, job_id="pending", tipo="cartao-ponto", status="processando")

    replacement = {
        "pages": [{"page": 1, "days": [{"date_raw": "02/01/2025", "punches": []}]}]
    }
    response = client.put("/api/transcricoes/done", json={"value": replacement})
    assert response.status_code == 200
    assert response.json()["value"] == replacement

    invalid = client.put("/api/transcricoes/done", json={"value": {"pages": "inválido"}})
    assert invalid.status_code == 422
    conflict = client.put("/api/transcricoes/pending", json={"value": replacement})
    assert conflict.status_code == 409


def test_put_replacement_is_used_by_subsequent_download(client, cartao_value):
    _insert_job(client, job_id="editable", tipo="cartao-ponto", status="concluido", value=cartao_value)
    corrected = {
        "pages": [
            {
                "page": 1,
                "days": [
                    {
                        "date_raw": "03/01/2025",
                        "punches": [
                            {"kind": "IN", "time_raw": "07:45", "time_hhmm": "07:45"},
                            {"kind": "OUT", "time_raw": "17:15", "time_hhmm": "17:15"},
                        ],
                    }
                ],
            }
        ]
    }
    assert client.put("/api/transcricoes/editable", json={"value": corrected}).status_code == 200

    download = client.get("/api/transcricoes/editable/planilha?formato=csv")
    assert download.status_code == 200
    rows = list(csv.reader(StringIO(download.content.decode("utf-8"))))
    assert rows == [["Data", "Entrada 1", "Saída 1"], ["03/01/2025", "07:45", "17:15"]]
    assert "8:00" not in download.text


def test_put_and_json_download_preserve_all_four_corrected_payroll_field_strings(client):
    original = {
        "pages": [
            {
                "page": 1,
                "year": "2025",
                "month": "01",
                "fields": [
                    {"code": "001", "label": "ALARIO", "reference": "220,00", "value": "1.000,00"}
                ],
                "bases": [],
            }
        ]
    }
    _insert_job(
        client,
        job_id="payroll-editable",
        tipo="holerite",
        status="concluido",
        value=original,
    )
    corrected = {
        "pages": [
            {
                **original["pages"][0],
                "fields": [
                    {
                        "code": "009",
                        "label": "SALÁRIO CORRIGIDO",
                        "reference": "200,00",
                        "value": "9.999,99",
                    }
                ],
            }
        ]
    }

    put = client.put("/api/transcricoes/payroll-editable", json={"value": corrected})
    assert put.status_code == 200
    assert put.json()["value"] == corrected
    assert client.get("/api/transcricoes/payroll-editable").json()["value"] == corrected
    downloaded = client.get(
        "/api/transcricoes/payroll-editable/planilha?formato=json"
    )
    assert downloaded.status_code == 200
    assert downloaded.json() == corrected


def test_download_conflicts_until_job_is_completed(client):
    assert client.get("/api/transcricoes/missing/planilha?formato=json").status_code == 404
    _insert_job(client, job_id="pending", tipo="holerite", status="processando")
    _insert_job(client, job_id="failed", tipo="holerite", status="erro", erro="Falha legível.")
    assert client.get("/api/transcricoes/pending/planilha?formato=json").status_code == 409
    assert client.get("/api/transcricoes/failed/planilha?formato=json").status_code == 409
    assert client.get("/api/transcricoes/pending/planilha?formato=pdf").status_code == 422


def test_corrected_repeated_payroll_field_reaches_json_csv_and_xlsx(client):
    original = {
        "pages": [
            {
                "page": 2,
                "year": "2017",
                "month": "12",
                "fields": [
                    {"code": "419", "label": "13º. Adto Desc", "reference": "", "value": "100,00"},
                    {"code": "419", "label": "13º. Adto Desc", "reference": "", "value": "200,00"},
                ],
                "bases": [],
            }
        ]
    }
    _insert_job(
        client,
        job_id="duplicate-fields",
        tipo="holerite",
        status="concluido",
        value=original,
    )
    corrected = {
        "pages": [
            {
                **original["pages"][0],
                "fields": [
                    original["pages"][0]["fields"][0],
                    {
                        **original["pages"][0]["fields"][1],
                        "value": "9.999,99",
                    },
                ],
            }
        ]
    }

    put = client.put("/api/transcricoes/duplicate-fields", json={"value": corrected})
    assert put.status_code == 200
    assert client.get("/api/transcricoes/duplicate-fields").json()["value"] == corrected

    json_response = client.get(
        "/api/transcricoes/duplicate-fields/planilha?formato=json"
    )
    assert json_response.status_code == 200
    assert json_response.json() == corrected

    csv_response = client.get(
        "/api/transcricoes/duplicate-fields/planilha?formato=csv"
    )
    assert csv_response.status_code == 200
    csv_rows = list(csv.reader(StringIO(csv_response.content.decode("utf-8"))))
    assert csv_rows == [
        ["Pág.", "Mês", "Ano", "13º. Adto Desc", "13º. Adto Desc (2)"],
        ["2", "12", "2017", "100,00", "9.999,99"],
    ]

    xlsx_response = client.get(
        "/api/transcricoes/duplicate-fields/planilha?formato=xlsx"
    )
    assert xlsx_response.status_code == 200
    worksheet = load_workbook(BytesIO(xlsx_response.content)).active
    assert [cell.value for cell in worksheet[1]] == [
        "Pág.",
        "Mês",
        "Ano",
        "13º. Adto Desc",
        "13º. Adto Desc (2)",
    ]
    assert worksheet.cell(row=2, column=4).value == "100,00"
    assert worksheet.cell(row=2, column=5).value == "9.999,99"


def test_healthz_checks_database(client):
    assert client.get("/healthz").json() == {"status": "ok"}
